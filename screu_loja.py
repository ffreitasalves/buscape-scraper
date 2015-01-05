import requests
import bs4
import openpyxl
import logging
from openpyxl import load_workbook
from openpyxl import Workbook

#iniciar o arquivo de log
logging.basicConfig(filename='arq_lojas.log',level=logging.WARNING)

#Abrindo o excel de leitura
wb = load_workbook(filename = 'lojas.xlsx', use_iterators = True)
ws = wb.get_sheet_by_name(name = 'Sheet1')

#iniciar a planilha do excel de excrita
wb_escrita = Workbook(write_only = True)
ws_escrita = wb_escrita.create_sheet()

for row in ws.iter_rows():
	try:
		nome = row[0].value
		url = row[1].value
		response = requests.get(url)
		soup = bs4.BeautifulSoup(response.text)

		div = soup.find(class_='company-header')
		ebit = ''
		if div.find('img'):
			if div.find('img').has_attr('title'):
				ebit = div.find('img')['title']

		
		desc = div.find(class_='about').string

		div = soup.find(class_='company-contacts')
		lista = []
		for li in div.find_all('li'):
			lista.append (li.text)

		ws_escrita.append(
			[nome,
			url,
			ebit,
			desc] + lista
			)
	except Exception as err:
		print "Erro na pagina %s" % url
		logging.exception(u"Erro na pagina %s" % url) 

wb_escrita.save('lojas_final.xlsx')